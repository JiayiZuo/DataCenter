#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多币种收付款模块 - 存证+Excel导入自动化脚本（最终版）
功能：
1. 调用存证接口获取txHash
2. 将原始txHash（含0x）直接作为资产编号写入Excel
3. 调用/inventory/importFile接口上传Excel
"""

import requests
import json
import openpyxl
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import uuid
import hashlib
from typing import List, Optional


# 配置区域
class Config:
    # 区块链存证服务配置
    EVIDENCE_BASE_URL = "http://172.22.152.154:8090"
    CHAIN_ID = "FELGN5IWTZB4"

    # 存证服务认证信息
    EVIDENCE_API_KEY = "058b6995c95e45a2bcd2f189a2334ad1"
    EVIDENCE_IDENTITY_ID = "FCAVLRLUFYTC"
    EVIDENCE_COOKIE = "Secure"

    # 库存导入服务配置
    INVENTORY_BASE_URL = "http://47.92.193.45:31880"
    INVENTORY_TOKEN= "eyJ0eXBlIjoiSldUIiwiYWxnIjoiSFMyNTYifQ.eyJyb2xlIjoiYWRtaW4iLCJ1c2VySWQiOiIxIiwic3ViIjoiYWRtaW4iLCJpYXQiOjE3NzI2MDU3OTEsImV4cCI6MTc3Nzc4OTc5MSwibmJmIjoxNzcyNjA1NzkxfQ.8ccQ-ciUlSXudjpA2_TXC-p0zT5lezowha37fAo2YDs"

    # Excel模板列配置
    EXCEL_COLUMNS = [
        "序号",
        "资产编号",  # 必填，使用原始txHash（含0x）
        "资产名称",
        "规格型号",
        "EPC",  # 必填，16进制，长度≤132，小写
    ]

    EPC_LENGTH = 64
    OUTPUT_DIR = "./output"


# 工具函数
class Utils:
    """工具函数类"""
    @staticmethod
    def generate_epc(tx_hash: str, index: int = 0) -> str:
        """
        生成符合规范的EPC编码
        规则：仅16进制字符，长度不超过132，全部小写，唯一性
        """
        epc_source = f"{tx_hash}{index}{datetime.now().timestamp()}"
        epc_hash = hashlib.sha256(epc_source.encode()).hexdigest()
        epc = epc_hash[:Config.EPC_LENGTH].lower()
        return epc

    @staticmethod
    def ensure_dir(directory: str):
        """确保目录存在"""
        if not os.path.exists(directory):
            os.makedirs(directory)

    @staticmethod
    def generate_filename(prefix: str = "asset_import") -> str:
        """生成带时间戳的文件名"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"

# 存证服务类
class EvidenceService:
    """区块链存证服务"""

    def __init__(self, base_url: str, chain_id: str, api_key: str,
                 identity_id: str, idempotency_key: str, cookie: str = None):
        self.base_url = base_url.rstrip('/')
        self.chain_id = chain_id
        self.api_key = api_key
        self.identity_id = identity_id
        self.idempotency_key = idempotency_key

        self.headers = {
            "Accept": "application/json",
            "Content-Type": "application/json;charset=UTF-8",
            "Authorization": f"Bearer {api_key}",
            "Identity-Id": identity_id,
            # "Idempotency-Key": str(uuid.uuid4())
        }

        if cookie:
            self.headers["Cookie"] = cookie

    def submit_evidence(self, biz_trace_id: str, record_data: dict, evidence_type: str = "TODO") -> Optional[dict]:
        """提交存证数据到区块链"""

        url = f"{self.base_url}/v1/evidence/chains/{self.chain_id}/records"

        payload = {
            "bizTraceId": biz_trace_id,
            "type": evidence_type,
            "record": record_data
        }

        headers = self.headers
        headers["Idempotency-Key"] = str(uuid.uuid4())

        try:
            response = requests.post(url, headers=self.headers, json=payload, timeout=30)
            if response.status_code == 401:
                return None

            response.raise_for_status()
            result = response.json()

            if result.get("code") == 200 or result.get("code") == "200":
                data = result.get("data", {})
                chain_tx_data = data.get("chainTxData", {})
                print(f"    - 记录ID: {data.get('id', 'N/A')}")
                print(f"    - txHash: {chain_tx_data.get('txHash', 'N/A')}")
                print(f"    - 区块高度：{chain_tx_data.get('blockNumber', 'N/A')}")

                return data
            else:
                return None

        except requests.exceptions.RequestException as e:
            return None
        except json.JSONDecodeError as e:
            return None

# Excel处理类
class ExcelHandler:
    @staticmethod
    def create_asset_excel(asset_list: List[dict],
                           output_file: str,
                           sheet_name: str = "资产清单列表") -> str:

        Utils.ensure_dir(os.path.dirname(output_file) if os.path.dirname(output_file) else ".")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        headers = Config.EXCEL_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 写入数据 todo asset_list数据
        for row_idx, asset in enumerate(asset_list, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=asset.get("asset_no", ""))
            ws.cell(row=row_idx, column=3, value=asset.get("asset_name", ""))
            ws.cell(row=row_idx, column=4, value=asset.get("spec_model", ""))
            epc = asset.get("epc", "")
            ws.cell(row=row_idx, column=5, value=epc.lower() if epc else "")
            ws.cell(row=row_idx, column=6, value=asset.get("position", ""))

        # 设置列宽
        column_widths = [8, 66, 25, 20, 40, 20]
        for col, width in enumerate(column_widths, 1):
            if col <= ws.max_column:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 添加边框
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(asset_list) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        wb.save(output_file)
        print(f"✓ Excel文件已生成：{output_file}")
        print(f"  - 记录数：{len(asset_list)}")
        return output_file

    @staticmethod
    def validate_excel(file_path: str) -> tuple:
        """验证Excel文件是否符合规范"""
        errors = []

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                row_data = dict(zip(headers, row))

                # 检查必填字段
                if not row_data.get("资产编号"):
                    errors.append(f"第{idx}行：资产编号不能为空")

                if not row_data.get("EPC"):
                    errors.append(f"第{idx}行：EPC不能为空")
                else:
                    epc = str(row_data.get("EPC", "")).lower()
                    if not all(c in '0123456789abcdef' for c in epc):
                        errors.append(f"第{idx}行：EPC必须为16进制字符")
                    if len(epc) > 132:
                        errors.append(f"第{idx}行：EPC长度不能超过132")

                # 检查资产编号格式（允许0x，不允许其他特殊标点符号）
                asset_no = str(row_data.get("资产编号", ""))
                special_chars = "!@#$%^&*()+=[]{}|;':\",./<>?"
                if any(c in asset_no for c in special_chars):
                    errors.append(f"第{idx}行：资产编号包含特殊字符")

            return (len(errors) == 0, errors)

        except Exception as e:
            return (False, [f"文件读取错误：{str(e)}"])

# 库存导入服务类
class InventoryService:
    def __init__(self, base_url: str, bearer_token: str):

        self.base_url = base_url.rstrip('/')
        self.bearer_token = bearer_token
        self.headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {bearer_token}"
        }

    def import_file(self, file_path: str) -> Optional[dict]:
        url = f"{self.base_url}/api/inventory/updateFile"

        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"✗ 文件不存在：{file_path}")
            return None

        try:
            with open(file_path, 'rb') as f:
                files = {
                    'file': (
                        os.path.basename(file_path),
                        f,
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                }

                # 调用接口
                response = requests.post(
                    url,
                    headers=self.headers,
                    files=files,
                    timeout=60
                )

                # 处理401认证错误
                if response.status_code == 401:
                    return None

                # 处理其他错误状态码
                if response.status_code >= 400:
                    try:
                        error_data = response.json()
                        print(f"  → 错误信息：{error_data}")
                    except:
                        print(f"  → 响应内容：{response.text[:500]}")
                    return None

                # 解析成功响应
                try:
                    result = response.json()
                    return result
                except json.JSONDecodeError:
                    return {"status": "SUCCESS", "statusCode": response.status_code}

        except requests.exceptions.ConnectionError as e:
            print(f"  ✗ 连接失败：无法连接到 {self.base_url}")
            print(f"     请确认服务是否已启动")
            print(f"     错误详情：{str(e)}")
            return None
        except requests.exceptions.Timeout as e:
            print(f"  ✗ 请求超时：{str(e)}")
            return None
        except requests.exceptions.RequestException as e:
            print(f"  ✗ 接口调用异常：{str(e)}")
            return None
        except Exception as e:
            print(f"  ✗ 未知错误：{str(e)}")
            return None

# 主流程类
class AssetImportWorkflow:
    def __init__(self):
        # 初始化存证服务
        self.evidence_service = EvidenceService(
            base_url=Config.EVIDENCE_BASE_URL,
            chain_id=Config.CHAIN_ID,
            api_key=Config.EVIDENCE_API_KEY,
            identity_id=Config.EVIDENCE_IDENTITY_ID,
            idempotency_key="",
            cookie=Config.EVIDENCE_COOKIE
        )

        self.inventory_service = InventoryService(
            base_url=Config.INVENTORY_BASE_URL,
            bearer_token=Config.INVENTORY_TOKEN
        )

        Utils.ensure_dir(Config.OUTPUT_DIR)

    def execute_batch(self, file_name: str) -> List[dict]:
        # 读取初始excel文件中信息
        current_dir = os.getcwd()
        if file_name:
            file_path = os.path.join(current_dir, file_name)
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在：{file_path}")
        else:
            # 自动查找当前目录下第一个 .xlsx 文件
            xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx') and not f.startswith('~')]
            if not xlsx_files:
                raise FileNotFoundError("当前目录下未找到 Excel 文件 (.xlsx)")
            file_name = xlsx_files[0]
            file_path = os.path.join(current_dir, file_name)

        wb = load_workbook(file_path, data_only=True)

        header_mapping = {
            "序号": "id",
            "资产编号": "asset_id",
            "资产名称": "asset_name",
            "规格型号": "spec_model",
            "EPC": "epc",
        }
        ws = wb.active

        asset_list = []
        headers = []
        col_mapping = {}  # {目标字段名：列索引}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # 跳过全空行
            if not any(cell is not None for cell in row):
                continue

            if row_idx == 1:
                # 第一行作为表头，建立映射
                for col_idx, cell_value in enumerate(row, 1):
                    if cell_value and cell_value in header_mapping:
                        target_field = header_mapping[cell_value]
                        col_mapping[target_field] = col_idx
                headers = [cell_value for cell_value in row if cell_value]
            else:
                # 数据行转换
                asset_data = {
                    "id": "",
                    "asset_id": "",
                    "asset_name": "",
                    "spec_model": "",
                    "epc": ""
                }

                # 根据映射填充数据
                for target_field, col_idx in col_mapping.items():
                    if col_idx <= len(row):
                        value = row[col_idx - 1]
                        if value is not None:
                            asset_data[target_field] = value


                asset_list.append(asset_data)

        wb.close()

        results = []
        evidence_results = []

        # 步骤1: 批量提交存证
        print("\n【步骤1】批量提交区块链存证")
        for idx, asset in enumerate(asset_list, 1):
            print(f"\n  处理资产 {idx}/{len(asset_list)}...")

            biz_trace_id = f"BATCH-{datetime.now().strftime('%Y%m%d')}-{idx:04d}"
            evidence_record = {
                "files": [
                    {
                        "name": asset.get("asset_name", f"asset_{idx}.img"),
                        "hash": hashlib.sha256(asset.get("asset_name", "").encode()).hexdigest(),
                        "time": int(datetime.now().timestamp())
                    }
                ]
            }

            result = self.evidence_service.submit_evidence(
                biz_trace_id=biz_trace_id,
                record_data=evidence_record,
                evidence_type="TODO 自定义"
            )
            evidence_results.append(result)

        # 步骤2: 生成Excel
        print("\n【步骤2】生成Excel文件")
        excel_data = []

        for idx, (asset, evidence) in enumerate(zip(asset_list, evidence_results)):
            if evidence:
                tx_hash = evidence.get("chainTxData", {}).get("txHash", "")
                asset_no = tx_hash
                epc = Utils.generate_epc(tx_hash, idx)

                excel_data.append({
                    "asset_no": asset_no,
                    "asset_name": asset.get("asset_name", ""),
                    "spec_model": asset.get("spec_model", ""),
                    "epc": epc,
                    "position": asset.get("position", "") if asset.get("is_all_in_one", False) else ""
                })

        if not excel_data:
            return results

        filename = Utils.generate_filename("batch_asset_import")
        output_file = os.path.join(Config.OUTPUT_DIR, filename)
        ExcelHandler.create_asset_excel(excel_data, output_file)

        # 步骤3: 调用库存导入接口
        print("\n【步骤3】调用库存导入服务/inventory/importFile接口")
        import_result = self.inventory_service.import_file(output_file)

        # 汇总结果
        for asset, evidence in zip(asset_list, evidence_results):
            if evidence:
                results.append({
                    "success": True,
                    "asset_name": asset.get("asset_name", ""),
                    "record_id": evidence.get("id", ""),
                    "tx_hash": evidence.get("chainTxData", {}).get("txHash", ""),
                    "excel_file": output_file
                })
            else:
                results.append({
                    "success": False,
                    "asset_name": asset.get("asset_name", ""),
                    "error": "Evidence submission failed"
                })
        return results

def main():
    load_dotenv()
    workflow = AssetImportWorkflow()
    file_name = "test.xlsx"

    results = workflow.execute_batch(file_name)
    return results


if __name__ == "__main__":
    main()